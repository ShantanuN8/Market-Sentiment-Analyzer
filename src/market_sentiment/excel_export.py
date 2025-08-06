from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

class ExcelExporter:
    def export(self, data,summary, filename='market_sentiment_report.xlsx'):
        wb=Workbook()
        ws=wb.active
        ws.title='Market Sentiment'

        headers=['Timestamp','Source','Headline','Textblob Score','VADER Score','Final Sentiment','Final Score','URL']
        ws.append(headers)
        for d in data:
            ws.append([d['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                       d['source'], d['headline'], round(d['textblob_score'],3),round(d['vader_score'],3),d['final_sentiment'],round(d['final_score'],3),d['url']])
        for cell in ws[1]:
            cell.font=Font(bold=True)
            cell.alignment=Alignment(horizontal='center',wrap_text=True)
        for col in ws.columns:
            max_length=0
            col_letter=col[0].column_letter   
            try:
                max_length=max(max_length,len(str(cell.value)))
            except:
                pass
            ws.column_dimensions[col_letter].width=min(max_length+5, 60)

        output_dir=os.path.join('outputs','reports')
        os.makedirs(output_dir,exist_ok=True)
        path=os.path.join(output_dir, filename)
        wb.save(path)
        return path
